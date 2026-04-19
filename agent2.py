import json
import os
from typing import List, Optional
import logging
import pandas as pd
from pydantic import BaseModel, Field
from llm_backend import llm

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger("Agent_2")


class AuditResult(BaseModel):
    is_complete: bool = Field(
        description="True, если найдены все необходимые компоненты"
    )
    missing_items: List[str] = Field(
        description="Список позиций из спецификации, паспорта которых не найдены"
    )
    extra_items: List[str] = Field(
        description="Список найденных паспортов, которых нет в спецификации (лишние)"
    )
    feedback_for_agent_1: str = Field(
        description="Подсказка для Агента 1, где и что поискать повторно, если чего-то не хватает"
    )


class CabinetAgent:
    def __init__(
        self, cabinet_name: str, cabinet_sn: str, storage_dir: str = "data/cabinets"
    ):
        self.cabinet_name = cabinet_name
        self.cabinet_sn = cabinet_sn
        self.storage_dir = storage_dir
        self.model_name = "qwen2.5:3b"
        self.file_path = os.path.join(
            storage_dir, f"{cabinet_name.replace(' ', '_')}.json"
        )

        os.makedirs(self.storage_dir, exist_ok=True)

        self.cabinet_data = {
            "name": self.cabinet_name,
            "serial_number": self.cabinet_sn,
            "expected_items": [],  # План
            "found_items": [],  # Факт
            "audit_report": None,  # Результат проверки
        }
        self._load_from_disk()

    def set_expected_plan(self, parsed_list: List[dict]):
        self.cabinet_data["expected_items"] = parsed_list
        self._save_to_disk()
        logger.info(f"План шкафа обновлен. Позиций {len(parsed_list)}")

    def add_found_passports(self, parsed_passports: List[dict]):
        self.cabinet_data["found_items"].extend(parsed_passports)
        self._save_to_disk()
        logger.info(f"Добавлено {len(parsed_passports)} паспортов")

    def _save_to_disk(self):
        with open(self.file_path, "w", encoding="utf-8") as f:
            json.dump(self.cabinet_data, f, ensure_ascii=False, indent=2)

    def _load_from_disk(self):
        if os.path.exists(self.file_path):
            with open(self.file_path, "r", encoding="utf-8") as f:
                self.cabinet_data = json.load(f)

    def run_audit(self) -> Optional[AuditResult]:
        logger.info("Запуск интеллектуальной сверки (Аудит)...")

        expected_names = [
            item.get("name", "") for item in self.cabinet_data["expected_items"]
        ]
        found_names = [
            item.get("name", "") for item in self.cabinet_data["found_items"]
        ]

        if not expected_names or not found_names:
            logger.warning("Нет данных для сверки.")
            return None

        prompt = f"""
        ОЖИДАЕМАЯ СПЕЦИФИКАЦИЯ: {json.dumps(expected_names, ensure_ascii=False)}
        ФАКТИЧЕСКИ НАЙДЕННЫЕ ПАСПОРТА: {json.dumps(found_names, ensure_ascii=False)}
        ЗАДАЧА: Сравни списки. Найди нехватку и лишние детали.
        """

        try:
            schema = AuditResult.model_json_schema()

            response = llm.create_chat_completion(
                messages=[
                    {
                        "role": "system",
                        "content": "Ты Инженер ОТК. Твоя задача — найти расхождения между планом и фактом.",
                    },
                    {"role": "user", "content": prompt},
                ],
                response_format={"type": "json_object", "schema": schema},
                temperature=0.1,
            )

            json_str = response["choices"][0]["message"]["content"]
            audit_result = AuditResult.model_validate_json(json_str)

            self.cabinet_data["audit_report"] = audit_result.model_dump()
            self._save_to_disk()
            return audit_result

        except Exception as e:
            logger.error(f"Ошибка LLM аудитора: {e}")
            return None

    def _print_report(self, result: AuditResult):
        if result.is_complete:
            print("\nШкаф укомплектован. Расхождений нет.")
        else:
            print("\nВНИМАНИЕ: Обнаружены расхождения!")
            if result.missing_items:
                print(f"Не хватает: {', '.join(result.missing_items)}")
            if result.extra_items:
                print(f"Лишние паспорта: {', '.join(result.extra_items)}")
            print(f"Рекомендация: {result.feedback_for_agent_1}")

    def tool_export_to_excel(self) -> str:
        items_to_export = self.cabinet_data.get("expected_items", [])

        if not items_to_export:
            logger.warning("Нет данных для экспорта в Excel.")
            return ""

        formatted_data = []
        for index, item in enumerate(items_to_export, start=1):
            doc_name = item.get("name", "Не указано")
            if item.get("article"):
                doc_name += f"\nАртикул: {item.get('article')}"

            sn_value = item.get("serial_number") or item.get("sn") or "б/н"

            formatted_data.append(
                {
                    "№ п/п": index,
                    "Наименование документа": doc_name,
                    "Заводской номер": sn_value,
                    "Страниц / листов": item.get("pages", "1 лист"),
                    "Сертификат": item.get("certificate", "-"),
                }
            )

        df = pd.DataFrame(formatted_data)
        export_path = os.path.join(
            self.storage_dir, f"Перечень_{self.cabinet_name.replace(' ', '_')}.xlsx"
        )

        with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Перечень")

            worksheet = writer.sheets["Перечень"]

            worksheet.column_dimensions["A"].width = 8  # № п/п
            worksheet.column_dimensions["B"].width = 60
            worksheet.column_dimensions["C"].width = 20  # Заводской номер
            worksheet.column_dimensions["D"].width = 18  # Страниц / листов
            worksheet.column_dimensions["E"].width = 15  # Сертификат

            from openpyxl.styles import Alignment

            for row in worksheet.iter_rows(min_row=2, max_col=5):
                row[1].alignment = Alignment(wrap_text=True, vertical="center")
                row[2].alignment = Alignment(horizontal="center", vertical="center")

        logger.info(f"Excel отчет сгенерирован по формату: {export_path}")
        return export_path

    def tool_get_qr_data_batch(self) -> List[dict]:
        qr_batch = []
        for item in self.cabinet_data["found_items"]:
            if item.get("serial_number") and item.get("serial_number") != "б/н":
                qr_batch.append(
                    {
                        "name": item.get("name"),
                        "sn": item.get("serial_number"),
                        "cabinet": self.cabinet_name,
                    }
                )
        logger.info(f"Подготовлено {len(qr_batch)} записей длля печати QR")
        return qr_batch

    def ask_rag_assistant(self, user_question: str) -> str:
        logger.info(f"RAG запрос: {user_question}")

        context_data = json.dumps(self.cabinet_data["found_items"], ensure_ascii=False)
        prompt = f"""
        Ответь на вопрос инженера, используя ТОЛЬКО данные из базы паспортов текущего шкафа.
        Если ответа в базе нет, честно скажи "В распознанных паспортах нет такой информации".
        
        БАЗА ПАСПОРТОВ:
        {context_data}
        
        ВОПРОС: {user_question}
        """

        try:
            response = llm.create_chat_completion(
                messages=[
                    {
                        "role": "system",
                        "content": "Ты — технический ассистент на производстве.",
                    },
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=1000,
            )
            return response["choices"][0]["message"]["content"]

        except Exception as e:
            logger.error(f"Ошибка RAG: {e}")
            return "Нет связи с моделью"

    def tool_export_passport_card(self, item: dict) -> str:
        from openpyxl.styles import Font, Alignment, Border, Side

        card_data = [
            {"Параметр": "ОСНОВНАЯ ИНФОРМАЦИЯ", "Значение": ""},
            {"Параметр": "Наименование", "Значение": item.get("name")},
            {"Параметр": "Артикул / Модель", "Значение": item.get("article") or "-"},
            {
                "Параметр": "Заводской номер",
                "Значение": item.get("serial_number") or item.get("sn") or "б/н",
            },
            {"Параметр": "Количество в шкафу", "Значение": item.get("quantity", 1)},
            {"Параметр": "", "Значение": ""},
            {
                "Параметр": "ТЕХНИЧЕСКИЕ ХАРАКТЕРИСТИКИ",
                "Значение": "",
            },
        ]
        specs = item.get("specifications", [])
        if specs:
            for s in specs:
                card_data.append(
                    {"Параметр": s.get("param_name"), "Значение": s.get("param_value")}
                )
        else:
            card_data.append({"Параметр": "Данные не найдены", "Значение": "-"})

        df = pd.DataFrame(card_data)

        safe_name = item.get("name", "Unknown").replace(" ", "_").replace("/", "_")
        filename = f"Карточка_{safe_name}.xlsx"
        export_path = os.path.join(self.storage_dir, filename)

        with pd.ExcelWriter(export_path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Паспорт")
            ws = writer.sheets["Паспорт"]

            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 55

            bold_font = Font(bold=True)
            for row in range(2, len(card_data) + 2):
                if ws.cell(row=row, column=2).value == "":
                    ws.cell(row=row, column=1).font = bold_font

            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="center", wrap_text=True)

        logger.info(f"Создана индивидуальная карточка: {export_path}")
        return export_path
